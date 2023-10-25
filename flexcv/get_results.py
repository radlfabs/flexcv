import pandas as pd
import neptune
import openpyxl
import numpy as np

from .model_mapping import MODEL_MAPPING

"""
Usage of this script:
Just add the desired run ids to the run_list in the bottom section of the module and run the script.
"""

api_dict = {
    "api_token": "ANONYMOUS",
    "project": "MYPROJECT",
}


def main(run_list: list[str], path="results_data.xlsx", exclude_nan=True):
    INDEX_NAMES = ("RunID", "Dataset", "Target", "ModelAcronym", "RS", "Fold", "Metric")
    # INDEX_NAMES = ("RunID", "Dataset", "Target", "ModelAcronym", "Fold", "Metric") # use this, when model is level 3
    L3_MODELS = (
        "LinearModel",
        "RandomForest",
        "XGBoost",
        "MARS",
        "SVR",
        "SVR_rbf",
    )  # TODO SVR
    L4_MODELS = (
        "LinearModel",
        "RandomForest",
        "XGBoost",
        "MARS",
        "SVR",
        "SVR_rbf",
        "MixedLM",
        "MERF",
        "XGBEM",
        "EarthEM",
        "SVREM",
    )
    METRICS = (
        "R2",
        "R2TRAIN",
        "MSE",
        "MSETRAIN",
        "MSE_IN",
        "MSETRAIN_IN",
        "RMSE",
        "RMSETRAIN",
        "MAE",
        "MAETRAIN",
    )

    run_dfs = []
    mapping_tracker = []
    # we want to visualize the following columns in the table:
    for run_id in run_list:
        url = f"https://app.neptune.ai/ISAVE-LLS/LLS-MethodsDatasetsComparison/e/{run_id}/metadata"
        print(f"\nFetching run {run_id}\n")
        # get the run using a context manager with a given run id
        with neptune.init_run(
            project="ISAVE-LLS/LLS-MethodsDatasetsComparison",
            with_id=run_id,
            api_token=api_dict["api_token"],
        ) as run:
            models_per_run = []
            # get the model acronym
            acronym = run["Acronym"].fetch()
            dataset = run["Dataset"].fetch()
            # onyl if we have RS, fetch RS
            try:
                random_slope = run["RS"].fetch()
                # INDEX_NAMES = ("RunID", "Dataset", "Target", "ModelAcronym", "RS", "Fold", "Metric") # use this, when model is level 4
            except (
                neptune.exceptions.MissingFieldException
            ):  # if we don't have RS, set it to some string
                random_slope = "noRS"
            target = run["AV"].fetch()
            description = run["sys/description"].fetch()
            mapping_tracker.append([run_id, url])

            # assign models corresponding to the acronym
            models = L3_MODELS if "3" in acronym else L4_MODELS

            # iterate over models to fetch the corresponding dataframe of values
            for model in models:
                # make a list to store both metrics
                model_dfs = []

                # iterate over metrics
                for metric in METRICS:
                    idx = f"{model}/{metric}"
                    try:
                        values_df = (
                            run[idx]
                            .fetch_values()  # API call to fetch values from the run
                            .astype({"step": "int8"})  # cast step to int8
                            .rename(
                                columns={"value": f"{model}"}
                            )  # rename the value column to the model name
                        )
                        # make a multiindex
                        values_df.index = pd.MultiIndex.from_tuples(
                            list(
                                zip(
                                    [
                                        run_id for _ in range(5)
                                    ],  # repeat the run id 5 times
                                    [
                                        dataset for _ in range(5)
                                    ],  # repeat the dataset 5 times
                                    [
                                        target for _ in range(5)
                                    ],  # repeat the target 5 times
                                    [
                                        acronym for _ in range(5)
                                    ],  # repeat the acronym 5 times
                                    [
                                        random_slope for _ in range(5)
                                    ],  # repeat the RS 5 times
                                    values_df.step.tolist(),  # get the fold values
                                    [
                                        metric for _ in range(5)
                                    ],  # repeat the description 5 times
                                )
                            ),
                            names=INDEX_NAMES,
                        )

                        values_df = values_df.drop(
                            columns=["step", "timestamp"]
                        )  # drop the step and timestamp columns

                        if values_df.shape[0] < 2:
                            # only use statistics if number of values is > 1 else write string "number of values too low"
                            mean_val, median_val, std_val = np.nan, np.nan, np.nan
                        else:
                            try:
                                mean_val = values_df.loc[
                                    run_id,
                                    dataset,
                                    target,
                                    acronym,
                                    random_slope,
                                    :,
                                    metric,
                                ].mean()
                            except:
                                mean_val = np.nan
                            try:
                                median_val = values_df.loc[
                                    run_id,
                                    dataset,
                                    target,
                                    acronym,
                                    random_slope,
                                    :,
                                    metric,
                                ].median()
                            except:
                                median_val = np.nan
                            try:
                                std_val = values_df.loc[
                                    run_id,
                                    dataset,
                                    target,
                                    acronym,
                                    random_slope,
                                    :,
                                    metric,
                                ].std()
                            except:
                                std_val = np.nan

                        # assign statistics to the dataframe
                        values_df.loc[
                            run_id,
                            dataset,
                            target,
                            acronym,
                            random_slope,
                            "mean",
                            metric,
                        ] = mean_val
                        values_df.loc[
                            run_id,
                            dataset,
                            target,
                            acronym,
                            random_slope,
                            "median",
                            metric,
                        ] = median_val
                        values_df.loc[
                            run_id,
                            dataset,
                            target,
                            acronym,
                            random_slope,
                            "std",
                            metric,
                        ] = std_val

                        model_dfs.append(
                            values_df
                        )  # append the df with values of single metric to df of the model

                    except neptune.exceptions.MissingFieldException:
                        print(
                            f"Missing field for {model=} and {metric=}, can't fetch! Skipping..."
                        )
                        # continue

                # model_dfs now has two dataframes, one for each metric
                # we concatenate the metrics and append the model to the list of models
                try:
                    models_per_run.append(pd.concat(model_dfs))
                except ValueError as e:
                    print("ValueError while Concatenating.")

            # except neptune.exceptions.MissingFieldException:
            #     print(f"Model {model} or metrics not found in run {run_id}. Skipping...")

            # description is set from last iteration and is now written in a single column
            models_per_run[-1]["description"] = description
            # run_dfs now has 5 or 10 dataframes, one for each model
            # we concatenate them and append the run to the list of runs
            run_df = pd.concat(models_per_run, axis=1)
            run_df.index.names = INDEX_NAMES
            run_df = run_df.sort_index(level="Fold")
            run_dfs.append(run_df)

    # concatenate all the dfs of the models
    data = pd.concat(run_dfs)

    if exclude_nan:
        description = data["description"].copy()
        # if rows contain only NaNs, drop them
        data = data.drop(columns=["description"]).dropna(axis=0, how="all")
        # left join description to data
        data = data.join(description, how="left")

    # set index names because they get lost in concatenation
    data.index.names = INDEX_NAMES
    print(data)
    # write to excel
    data.to_excel(path, sheet_name="Results", merge_cells=False)

    # open the excel file again and make a new sheet to log the mapping of acronym to run
    wb = openpyxl.load_workbook(path)
    ws = wb.create_sheet("AcronymMapping")
    for row in mapping_tracker:
        ws.append(row)
    wb.save(path)
    # stop neptune run
    run.stop()
    return data


if __name__ == "__main__":
    run_list = [
        # "LLSUP-951",
        "LLSUP-952",
        "LLSUP-953",
        "LLSUP-951",
        "LLSUP-955",
        "LLSUP-956",
        "LLSUP-957",
    ]
    main(run_list, path="results_data.xlsx")
